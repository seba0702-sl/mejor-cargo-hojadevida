import streamlit as st
import pandas as pd
import re
from io import BytesIO
from dateutil.relativedelta import relativedelta
from components.charts import (
    preparar_rango_36_meses,
    generar_grafico_interactivo_secuencia,
)

st.set_page_config(page_title="Ir Datos de periodos 36 meses", layout="wide")

st.title("Grafico - 36 meses sin cortes")

GRUPOS_CARGO = [
    "",
    "MEJOR CARGO",
    "SIMULTANEO_1",
    "SIMULTANEO_2",
    "SIMULTANEO_3",
    "SIMULTANEO_4",
    "SIMULTANEO_5",
]


def valor_excel_seguro(value):
    if value is None:
        return ""

    if isinstance(value, (list, dict, set, tuple)):
        return str(value)

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, pd.Timestamp):
        return value.strftime("%d/%m/%Y")

    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass

    return value

def calcular_porcentaje_mejor_cargo():
    datos = st.session_state.get("datos_personales", {})
    fecha_nacimiento = datos.get("Fecha de nacimiento")
    fecha_cese = datos.get("Fecha de cese")

    if not fecha_nacimiento or not fecha_cese:
        return None

    fecha_nacimiento = pd.to_datetime(fecha_nacimiento)
    fecha_cese = pd.to_datetime(fecha_cese)

    df_serv = st.session_state.get("df_consolidado")

    if df_serv is None or df_serv.empty:
        return None

    df_serv = df_serv.copy()
    df_serv["DESDE"] = pd.to_datetime(df_serv["DESDE"], errors="coerce", dayfirst=True)
    df_serv["HASTA"] = pd.to_datetime(df_serv["HASTA"], errors="coerce", dayfirst=True)
    df_serv = df_serv.dropna(subset=["DESDE", "HASTA"]).sort_values("DESDE")

    if df_serv.empty:
        return None

    periodos = sorted(list(zip(df_serv["DESDE"], df_serv["HASTA"])), key=lambda x: x[0])

    unidos = []
    inicio, fin = periodos[0]

    for desde, hasta in periodos[1:]:
        if desde <= fin + pd.Timedelta(days=1):
            fin = max(fin, hasta)
        else:
            unidos.append((inicio, fin))
            inicio, fin = desde, hasta

    unidos.append((inicio, fin))

    total_dias = sum((fin - inicio).days for inicio, fin in unidos)

    base = pd.Timestamp("2000-01-01")
    servicios = relativedelta(base + pd.Timedelta(days=total_dias), base)
    edad = relativedelta(fecha_cese, fecha_nacimiento)

    info_anses = st.session_state.get("info_anses", {})
    tiene_anses = str(info_anses.get("tiene_anses", "No")).lower() in [
        "si",
        "sí",
        "true",
    ]

    otros_servicios_prorrateo = st.session_state.get(
        "otros_servicios_prorrateo",
        tiene_anses,
    )

    cumple_edad_70 = edad.years >= 50
    cumple_servicios_70 = servicios.years >= 25

    if edad.years >= 55 and servicios.years >= 30:
        return 80.0

    if edad.years >= 53 and servicios.years >= 28:
        return 75.0

    if cumple_edad_70 and (cumple_servicios_70 or otros_servicios_prorrateo):
        return 70.0

    return 0.0


def extraer_carga_numerica(valor):
    if pd.isna(valor):
        return 0.0

    match = re.search(r"(\d+(?:[.,]\d+)?)", str(valor))

    if not match:
        return 0.0

    return float(match.group(1).replace(",", "."))


def detectar_tipo_carga_grupo(row):
    texto = (
        str(row.get("CARGO", "")) + " " +
        str(row.get("ESCUELA", ""))
    ).upper()

    if any(p in texto for p in [
        "MODULO",
        "MODULOS",
        "MOD.",
        "HORAS MODULO",
        "MÓDULO",
        "MÓDULOS",
    ]):
        return "Modulos"

    if any(p in texto for p in [
        "HORAS CATEDRA",
        "HORAS CÁTEDRA",
        "HORAS DE CATEDRA",
        "HORAS DE CÁTEDRA",
        "HS CAT",
        "HS. CAT",
        "H/C",
        "PROFESOR HORAS",
        "PROFESOR HORA",
        "INSTRUCTOR",
    ]):
        return "Horas catedra"

    return "Cargo"


def calcular_totales_carga_grupo(df_calc, grupo):
    df_grupo = df_calc[df_calc["GRUPO"] == grupo].copy()

    if df_grupo.empty:
        return {
            "TOTAL_CARGOS_GRUPO": 0,
            "TOTAL_MODULOS_GRUPO": 0.0,
            "TOTAL_HORAS_CATEDRA_GRUPO": 0.0,
            "DETALLE_CARGA_GRUPO": "",
            "CARGO_RESUMEN_GRUPO": "",
        }

    df_grupo["TIPO_CARGA_GRUPO"] = df_grupo.apply(detectar_tipo_carga_grupo, axis=1)
    df_grupo["CARGA_NUMERICA_GRUPO"] = df_grupo["CARGA HORARIA"].apply(extraer_carga_numerica)

    df_unico = (
        df_grupo
        .sort_values("CARGA_NUMERICA_GRUPO")
        .groupby("SECUENCIA", as_index=False)
        .first()
    )

    total_cargos = df_unico["SECUENCIA"].astype(str).nunique()

    total_modulos = df_unico.loc[
        df_unico["TIPO_CARGA_GRUPO"] == "Modulos",
        "CARGA_NUMERICA_GRUPO",
    ].sum()

    total_horas_catedra = df_unico.loc[
        df_unico["TIPO_CARGA_GRUPO"] == "Horas catedra",
        "CARGA_NUMERICA_GRUPO",
    ].sum()

    cargos = " / ".join(sorted(set(df_unico["CARGO"].fillna("").astype(str))))

    detalle = (
        f"Cargos: {total_cargos} | "
        f"Modulos: {total_modulos:g} | "
        f"Horas catedra: {total_horas_catedra:g}"
    )

    return {
        "TOTAL_CARGOS_GRUPO": int(total_cargos),
        "TOTAL_MODULOS_GRUPO": float(total_modulos),
        "TOTAL_HORAS_CATEDRA_GRUPO": float(total_horas_catedra),
        "DETALLE_CARGA_GRUPO": detalle,
        "CARGO_RESUMEN_GRUPO": cargos,
    }


def agregar_calculos_grupo(df_base):
    df_calc = df_base.copy()

    if "GRUPO" not in df_calc.columns:
        df_calc.insert(0, "GRUPO", "")

    df_calc["GRUPO"] = (
        df_calc["GRUPO"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df_calc["PORCENTAJE_GRUPO"] = pd.Series([pd.NA] * len(df_calc), dtype="object")
    df_calc["DESDE_MAX_GRUPO"] = pd.Series([""] * len(df_calc), dtype="object")
    df_calc["HASTA_MIN_GRUPO"] = pd.Series([""] * len(df_calc), dtype="object")
    df_calc["TOTAL_CARGOS_GRUPO"] = 0
    df_calc["TOTAL_MODULOS_GRUPO"] = 0.0
    df_calc["TOTAL_HORAS_CATEDRA_GRUPO"] = 0.0
    df_calc["DETALLE_CARGA_GRUPO"] = ""
    df_calc["CARGO_RESUMEN_GRUPO"] = ""

    porcentaje_mejor_cargo = calcular_porcentaje_mejor_cargo()

    for grupo in df_calc["GRUPO"].dropna().unique():
        if grupo == "":
            continue

        mask = df_calc["GRUPO"] == grupo
        totales = calcular_totales_carga_grupo(df_calc, grupo)

        df_calc.loc[mask, "TOTAL_CARGOS_GRUPO"] = totales["TOTAL_CARGOS_GRUPO"]
        df_calc.loc[mask, "TOTAL_MODULOS_GRUPO"] = totales["TOTAL_MODULOS_GRUPO"]
        df_calc.loc[mask, "TOTAL_HORAS_CATEDRA_GRUPO"] = totales["TOTAL_HORAS_CATEDRA_GRUPO"]
        df_calc.loc[mask, "DETALLE_CARGA_GRUPO"] = totales["DETALLE_CARGA_GRUPO"]
        df_calc.loc[mask, "CARGO_RESUMEN_GRUPO"] = totales["CARGO_RESUMEN_GRUPO"]

        if grupo == "MEJOR CARGO":
            if porcentaje_mejor_cargo is not None:
                df_calc.loc[mask, "PORCENTAJE_GRUPO"] = float(porcentaje_mejor_cargo)
            continue

        if grupo.startswith("SIMULTANEO_"):
            if "DESDE_REAL" not in df_calc.columns or "HASTA_REAL" not in df_calc.columns:
                df_calc.loc[mask, "PORCENTAJE_GRUPO"] = 0.0
                continue

            desde_real = pd.to_datetime(
                df_calc.loc[mask, "DESDE_REAL"],
                errors="coerce",
                dayfirst=True,
            )

            hasta_real = pd.to_datetime(
                df_calc.loc[mask, "HASTA_REAL"],
                errors="coerce",
                dayfirst=True,
            )

            desde_max_grupo = desde_real.max()
            hasta_min_grupo = hasta_real.min()

            if (
                pd.notna(desde_max_grupo)
                and pd.notna(hasta_min_grupo)
                and hasta_min_grupo >= desde_max_grupo
            ):
                dias = (hasta_min_grupo - desde_max_grupo).days
                anios = dias / 365.25
                porcentaje = round(anios * 2.8, 2)

                df_calc.loc[mask, "PORCENTAJE_GRUPO"] = float(porcentaje)
                df_calc.loc[mask, "DESDE_MAX_GRUPO"] = desde_max_grupo.strftime("%d/%m/%Y")
                df_calc.loc[mask, "HASTA_MIN_GRUPO"] = hasta_min_grupo.strftime("%d/%m/%Y")
            else:
                df_calc.loc[mask, "PORCENTAJE_GRUPO"] = 0.0

    return df_calc


def detectar_nivel_excel(escuela):
    niveles_dict_export = {
        "Inicial": ["JI", "JS", "JU", "JM", "JV"],
        "Primaria": ["PP", "EP", "PA", "DA", "DC", "DE"],
        "Especial": ["EE", "EL", "CFI", "ET", "ESPECIAL"],
        "Secundaria": ["MM", "MS", "ES", "ESB", "MT", "BS", "MA", "MC", "AS"],
        "Adulto": ["DM", "CENS", "DF", "DS", "MF", "ADULTOS", "ADULTO", "CFP", "CFL"],
        "Superior": ["IS", "AA", "AT", "AF", "AV", "AC", "AD", "AM", "AP", "FC"],
    }

    if pd.isna(escuela):
        return "Sin Nivel"

    escuela = str(escuela).upper()

    for nivel, codigos in niveles_dict_export.items():
        if any(codigo in escuela for codigo in codigos):
            return nivel

    return "Sin Nivel"


if "df_consolidado" not in st.session_state:
    st.warning("Primero carga y procesa los datos en la pagina principal.")
    st.stop()

df_consolidado = st.session_state["df_consolidado"]
hoja_licencias = st.session_state.get("hoja_licencias")
codigos_validos = st.session_state.get("codigos_validos")

df_base = df_consolidado.copy()

if "CONSOLIDADO" in df_base.columns:
    df_base = df_base[df_base["CONSOLIDADO"] == True].copy()

if df_base.empty:
    st.warning("No hay periodos consolidados para mostrar.")
    st.stop()

df_base["DESDE"] = pd.to_datetime(df_base["DESDE"], errors="coerce", dayfirst=True)
df_base["HASTA"] = pd.to_datetime(df_base["HASTA"], errors="coerce", dayfirst=True)
df_base = df_base.dropna(subset=["DESDE", "HASTA"])

if df_base.empty:
    st.warning("No hay fechas validas para trabajar.")
    st.stop()

if "fecha_control_consol" not in st.session_state:
    st.session_state.fecha_control_consol = df_base["DESDE"].min()

with st.sidebar:
    st.header("Antiguedad total")

    df_ant = df_consolidado.copy()
    df_ant["DESDE"] = pd.to_datetime(df_ant["DESDE"], errors="coerce", dayfirst=True)
    df_ant["HASTA"] = pd.to_datetime(df_ant["HASTA"], errors="coerce", dayfirst=True)
    df_ant["HASTA"] = df_ant["HASTA"].fillna(pd.Timestamp.today().normalize())
    df_ant = df_ant.dropna(subset=["DESDE", "HASTA"]).sort_values("DESDE")

    merged = []

    if not df_ant.empty:
        current_start = df_ant.iloc[0]["DESDE"]
        current_end = df_ant.iloc[0]["HASTA"]

        for _, row in df_ant.iloc[1:].iterrows():
            if row["DESDE"] <= current_end:
                current_end = max(current_end, row["HASTA"])
            else:
                merged.append((current_start, current_end))
                current_start, current_end = row["DESDE"], row["HASTA"]

        merged.append((current_start, current_end))

    total_days = sum((fin - inicio).days for inicio, fin in merged)

    base = pd.Timestamp("2000-01-01")
    rd = relativedelta(base + pd.Timedelta(days=total_days), base)

    st.success(
        f"Antiguedad total:\n\n"
        f"{rd.years} anos, {rd.months} meses, {rd.days} dias"
    )

st.markdown("### Navegar por fechas")
st.markdown("Lineas rojas: licencias")
st.markdown("Zona verde: ventana de 36 meses")

fechas_unicas = sorted(df_base["DESDE"].unique())
opciones_fechas = ["Todas las secuencias"] + list(fechas_unicas)

if st.session_state.fecha_control_consol not in fechas_unicas:
    st.session_state.fecha_control_consol = fechas_unicas[0]

fecha_select = st.selectbox(
    "Seleccionar fecha desde:",
    options=opciones_fechas,
    index=1,
    format_func=lambda x: x if isinstance(x, str) else x.strftime("%d/%m/%Y"),
)

st.session_state.fecha_control_consol = fecha_select

if fecha_select == "Todas las secuencias":
    st.markdown("### Mostrando todas las secuencias")
else:
    st.markdown(f"### Fecha seleccionada: {fecha_select.strftime('%d/%m/%Y')}")

if fecha_select == "Todas las secuencias":
    df_rango_36 = df_base.copy()
    fecha_inicio = None
    fecha_fin = None
else:
    fecha_inicio = fecha_select
    df_rango_36, fecha_fin = preparar_rango_36_meses(df_base, fecha_inicio)

fig = generar_grafico_interactivo_secuencia(
    df_rango_36,
    hoja_licencias,
    codigos_validos,
    df_rango_36["SECUENCIA"].astype(str).unique().tolist(),
)

if fecha_select != "Todas las secuencias":
    fig.add_shape(
        type="line",
        x0=fecha_inicio,
        x1=fecha_inicio,
        y0=0,
        y1=1,
        xref="x",
        yref="paper",
        line=dict(color="blue", width=2),
    )

    fig.add_shape(
        type="line",
        x0=fecha_fin,
        x1=fecha_fin,
        y0=0,
        y1=1,
        xref="x",
        yref="paper",
        line=dict(color="black", width=2),
    )

    fig.add_shape(
        type="rect",
        x0=fecha_inicio,
        x1=fecha_fin,
        y0=0,
        y1=1,
        xref="x",
        yref="paper",
        line=dict(width=0),
        fillcolor="rgba(0,255,0,0.35)",
        layer="above",
    )

st.plotly_chart(fig, use_container_width=True, key="gantt_36_meses")

st.markdown("### Secuencias validas (36 meses sin cortes)")

df_tabla = df_rango_36.copy()
df_tabla["DESDE"] = pd.to_datetime(df_tabla["DESDE"], errors="coerce", dayfirst=True)
df_tabla["HASTA"] = pd.to_datetime(df_tabla["HASTA"], errors="coerce", dayfirst=True)

secuencias_validas = []

for secuencia, grupo in df_tabla.groupby("SECUENCIA"):
    grupo = grupo.sort_values("DESDE")

    desde_min = grupo["DESDE"].min()
    hasta_max = grupo["HASTA"].max()

    if pd.notna(desde_min) and pd.notna(hasta_max) and fecha_inicio and fecha_fin:
        cubre_inicio = desde_min <= fecha_inicio
        cubre_fin = hasta_max >= fecha_fin
    else:
        cubre_inicio = False
        cubre_fin = False

    sin_cortes = True
    ultimo_hasta = None

    for _, row in grupo.iterrows():
        if ultimo_hasta is not None:
            if row["DESDE"] > ultimo_hasta + pd.Timedelta(days=1):
                sin_cortes = False
                break

        if ultimo_hasta is None:
            ultimo_hasta = row["HASTA"]
        else:
            ultimo_hasta = max(ultimo_hasta, row["HASTA"])

    if cubre_inicio and cubre_fin and sin_cortes:
        secuencias_validas.append(secuencia)

df_tabla = df_tabla[df_tabla["SECUENCIA"].isin(secuencias_validas)].copy()

if fecha_inicio is not None and fecha_fin is not None and not df_tabla.empty:
    df_tabla["DESDE_REAL"] = df_tabla["DESDE"]
    df_tabla["HASTA_REAL"] = df_tabla["HASTA"]
    df_tabla["DESDE_36_MESES"] = df_tabla["DESDE"].apply(lambda x: max(x, fecha_inicio))
    df_tabla["HASTA_36_MESES"] = df_tabla["HASTA"].apply(lambda x: min(x, fecha_fin))

df_tabla = df_tabla.sort_values(["SECUENCIA", "DESDE"]).reset_index(drop=True)

columnas_orden = [
    "SECUENCIA",
    "NIVEL" if "NIVEL" in df_tabla.columns else None,
    "DESDE_REAL",
    "HASTA_REAL",
    "DESDE_36_MESES",
    "HASTA_36_MESES",
    "ESCUELA",
    "CARGA HORARIA",
    "CARGO",
]

columnas_orden = [c for c in columnas_orden if c in df_tabla.columns]

if df_tabla.empty:
    st.warning("No hay secuencias que cubran los 36 meses completos sin cortes.")
else:
    st.success(f"{len(secuencias_validas)} secuencias cubren los 36 meses sin cortes")

    df_editor = df_tabla[columnas_orden].copy()

    if st.session_state.pop("limpiar_grupos_editor", False):
        st.session_state.pop("editor_secuencias_validas", None)
        st.session_state["df_secuencias_validas_grupos"] = pd.DataFrame()

    if "GRUPO" not in df_editor.columns:
        df_editor.insert(0, "GRUPO", "")

    st.info(
        "La columna Grupo editable se puede modificar. Hace clic en una celda "
        "para elegir MEJOR CARGO o un grupo SIMULTANEO."
    )

    if st.button("Limpiar grupos"):
        st.session_state["limpiar_grupos_editor"] = True
        st.session_state["df_secuencias_validas_grupos"] = pd.DataFrame()
        st.rerun()

    df_editor = st.data_editor(
        df_editor,
        use_container_width=True,
        height=450,
        hide_index=True,
        column_config={
            "GRUPO": st.column_config.SelectboxColumn(
                "Grupo editable",
                options=GRUPOS_CARGO,
                required=False,
                help="Elegi MEJOR CARGO o SIMULTANEO_1 a SIMULTANEO_5",
            )
        },
        disabled=[col for col in df_editor.columns if col != "GRUPO"],
        key="editor_secuencias_validas",
    )

    df_editor = agregar_calculos_grupo(df_editor)
    st.session_state["df_secuencias_validas_grupos"] = df_editor

    st.markdown("### Vista con grupos y porcentajes")

    st.dataframe(
        df_editor,
        use_container_width=True,
        height=300,
        hide_index=True,
    )

st.caption(
    "La tabla muestra los periodos reales y tambien el tramo que intersecta "
    "la ventana de 36 meses seleccionada."
)

if fecha_inicio is None or fecha_fin is None:
    st.info("Para exportar el analisis de 36 meses debes seleccionar una fecha.")
    st.stop()

buffer = BytesIO()

df_export = df_rango_36.copy()
df_export["DESDE"] = pd.to_datetime(df_export["DESDE"], errors="coerce", dayfirst=True)
df_export["HASTA"] = pd.to_datetime(df_export["HASTA"], errors="coerce", dayfirst=True)

fecha_desde_max = fecha_inicio
fecha_hasta_min = fecha_fin

df_filtrado = df_export[
    (df_export["HASTA"] >= fecha_desde_max)
    & (df_export["DESDE"] <= fecha_hasta_min)
].copy()

if df_filtrado.empty:
    st.warning("No hay datos dentro del rango visible de 36 meses.")
    st.stop()

secuencias_validas = []

for secuencia, grupo in df_filtrado.groupby("SECUENCIA"):
    grupo = grupo.sort_values("DESDE").copy()

    cursor = fecha_desde_max
    continuo = True

    for _, row in grupo.iterrows():
        desde = row["DESDE"]
        hasta = row["HASTA"]

        if pd.isna(desde) or pd.isna(hasta):
            continue

        desde = max(desde, fecha_desde_max)
        hasta = min(hasta, fecha_hasta_min)

        if hasta <= fecha_desde_max or desde >= fecha_hasta_min:
            continue

        if desde > cursor + pd.Timedelta(days=1):
            continuo = False
            break

        cursor = max(cursor, hasta)

        if cursor >= fecha_hasta_min:
            break

    if continuo and cursor >= fecha_hasta_min:
        secuencias_validas.append(secuencia)

df_export = df_export[df_export["SECUENCIA"].isin(secuencias_validas)].copy()

if df_export.empty:
    st.warning("Ninguna secuencia cubre los 36 meses completos sin cortes.")
    st.stop()

df_export["DESDE_REAL"] = df_export["DESDE"]
df_export["HASTA_REAL"] = df_export["HASTA"]
df_export["DESDE_36_MESES"] = df_export["DESDE"].apply(lambda x: max(x, fecha_desde_max))
df_export["HASTA_36_MESES"] = df_export["HASTA"].apply(lambda x: min(x, fecha_hasta_min))
df_export["DESDE_MAX_INTERSECCION"] = fecha_desde_max
df_export["HASTA_MIN_INTERSECCION"] = fecha_hasta_min

for col in [
    "DESDE_REAL",
    "HASTA_REAL",
    "DESDE_36_MESES",
    "HASTA_36_MESES",
    "DESDE_MAX_INTERSECCION",
    "HASTA_MIN_INTERSECCION",
]:
    df_export[col] = pd.to_datetime(df_export[col]).dt.strftime("%d/%m/%Y")

columnas_exportar = [
    "SECUENCIA",
    "CARGO",
    "DESDE_REAL",
    "HASTA_REAL",
    "DESDE_36_MESES",
    "HASTA_36_MESES",
    "ESCUELA",
    "CARGA HORARIA",
]

columnas_exportar = [c for c in columnas_exportar if c in df_export.columns]
df_export = df_export[columnas_exportar].copy()

df_grupos = st.session_state.get("df_secuencias_validas_grupos")

if df_grupos is not None and not df_grupos.empty:
    df_grupos_export = df_grupos.copy()

    for col in ["DESDE_REAL", "HASTA_REAL"]:
        if col in df_grupos_export.columns:
            df_grupos_export[col] = pd.to_datetime(
                df_grupos_export[col],
                errors="coerce",
                dayfirst=True,
            ).dt.strftime("%d/%m/%Y")

    columnas_grupo = [
        "SECUENCIA",
        "CARGO",
        "ESCUELA",
        "DESDE_REAL",
        "HASTA_REAL",
        "GRUPO",
        "PORCENTAJE_GRUPO",
        "DESDE_MAX_GRUPO",
        "HASTA_MIN_GRUPO",
        "TOTAL_CARGOS_GRUPO",
        "TOTAL_MODULOS_GRUPO",
        "TOTAL_HORAS_CATEDRA_GRUPO",
        "DETALLE_CARGA_GRUPO",
        "CARGO_RESUMEN_GRUPO",
    ]

    columnas_grupo = [c for c in columnas_grupo if c in df_grupos_export.columns]
    df_grupos_export = df_grupos_export[columnas_grupo].copy()

    for col in ["SECUENCIA", "CARGO", "ESCUELA", "DESDE_REAL", "HASTA_REAL"]:
        if col in df_grupos_export.columns and col in df_export.columns:
            df_grupos_export[col] = df_grupos_export[col].astype(str)
            df_export[col] = df_export[col].astype(str)

    df_export = df_export.merge(
        df_grupos_export,
        on=["SECUENCIA", "CARGO", "ESCUELA", "DESDE_REAL", "HASTA_REAL"],
        how="left",
    )
else:
    df_export["GRUPO"] = ""
    df_export["PORCENTAJE_GRUPO"] = ""
    df_export["DESDE_MAX_GRUPO"] = ""
    df_export["HASTA_MIN_GRUPO"] = ""
    df_export["TOTAL_CARGOS_GRUPO"] = 0
    df_export["TOTAL_MODULOS_GRUPO"] = 0.0
    df_export["TOTAL_HORAS_CATEDRA_GRUPO"] = 0.0
    df_export["DETALLE_CARGA_GRUPO"] = ""
    df_export["CARGO_RESUMEN_GRUPO"] = ""

df_export_resumen = df_export.copy()

columnas_exportar_final = [
    "GRUPO",
    "SECUENCIA",
    "CARGO",
    "DESDE_REAL",
    "HASTA_REAL",
    "DESDE_36_MESES",
    "HASTA_36_MESES",
    "ESCUELA",
    "CARGA HORARIA",
    "PORCENTAJE_GRUPO",
    "DESDE_MAX_GRUPO",
    "HASTA_MIN_GRUPO",
]

columnas_exportar_final = [c for c in columnas_exportar_final if c in df_export.columns]
df_export = df_export[columnas_exportar_final].copy()

df_resumen_grupo = df_export_resumen.copy()

if "GRUPO" in df_resumen_grupo.columns:
    df_resumen_grupo["GRUPO"] = df_resumen_grupo["GRUPO"].fillna("").astype(str)
    df_resumen_grupo = df_resumen_grupo[df_resumen_grupo["GRUPO"] != ""].copy()
else:
    df_resumen_grupo = pd.DataFrame()

if not df_resumen_grupo.empty:
    df_resumen_grupo["PORCENTAJE_GRUPO_NUM"] = pd.to_numeric(
        df_resumen_grupo["PORCENTAJE_GRUPO"],
        errors="coerce",
    ).fillna(0)

    df_resumen_grupo["TOTAL_MODULOS_GRUPO"] = pd.to_numeric(
        df_resumen_grupo["TOTAL_MODULOS_GRUPO"],
        errors="coerce",
    ).fillna(0)

    df_resumen_grupo["TOTAL_HORAS_CATEDRA_GRUPO"] = pd.to_numeric(
        df_resumen_grupo["TOTAL_HORAS_CATEDRA_GRUPO"],
        errors="coerce",
    ).fillna(0)

    df_resumen_pivot = (
        df_resumen_grupo
        .groupby("GRUPO", as_index=False)
        .agg(
            PORCENTAJE=("PORCENTAJE_GRUPO_NUM", "max"),
            CANTIDAD_MODULOS=("TOTAL_MODULOS_GRUPO", "max"),
            CANTIDAD_HS_CAT=("TOTAL_HORAS_CATEDRA_GRUPO", "max"),
            CARGO=("CARGO_RESUMEN_GRUPO", "first"),
        )
        .sort_values("GRUPO")
    )
else:
    df_resumen_pivot = pd.DataFrame(
        columns=[
            "GRUPO",
            "PORCENTAJE",
            "CANTIDAD_MODULOS",
            "CANTIDAD_HS_CAT",
            "CARGO",
        ]
    )

datos_personales = st.session_state.get("datos_personales", {})
nombre = datos_personales.get("Nombre y apellido", "")
edad = datos_personales.get("Edad", "")
cuil = datos_personales.get("CUIL", "")
clave_abc = datos_personales.get("Clave ABC", "")

info_anses = st.session_state.get("info_anses", {})
tiene_anses = info_anses.get("tiene_anses", "No")
servicios_anses = info_anses.get("tipos_servicio", [])
advertencia_anses = info_anses.get("advertencia_simultaneidad", "")

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Rango visible 36 meses"

ws["A1"] = "Nombre y apellido:"
ws["B1"] = nombre
ws["A2"] = "Edad:"
ws["B2"] = edad
ws["A3"] = "CUIL:"
ws["B3"] = cuil
ws["A4"] = "Clave ABC:"
ws["B4"] = clave_abc
ws["A5"] = "Tiene ANSES:"
ws["B5"] = tiene_anses
ws["A6"] = "Servicios ANSES:"
ws["B6"] = ", ".join(servicios_anses) if servicios_anses else ""
ws["A7"] = "Advertencia:"
ws["B7"] = advertencia_anses

start_row = 9

for r_idx, row in enumerate(
    dataframe_to_rows(df_export, index=False, header=True),
    start=start_row,
):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=valor_excel_seguro(value))

colores_nivel = {
    "Inicial": "FFC000",
    "Primaria": "92D050",
    "Especial": "B4C6E7",
    "Secundaria": "00B0F0",
    "Adulto": "F4B183",
    "Superior": "A9D18E",
    "Sin Nivel": "D9D9D9",
}

fila_inicio_datos = start_row + 1

for fila_excel, (_, row_df) in enumerate(
    df_export.iterrows(),
    start=fila_inicio_datos,
):
    nivel = detectar_nivel_excel(row_df["ESCUELA"])
    color = colores_nivel.get(nivel, "FFFFFF")

    fill = PatternFill(
        fill_type="solid",
        fgColor=color,
        start_color=color,
        end_color=color,
    )

    for col_excel in range(1, len(df_export.columns) + 1):
        ws.cell(row=fila_excel, column=col_excel).fill = fill

fila_resumen = ws.max_row + 3

ws.cell(row=fila_resumen, column=1, value="RESUMEN POR GRUPO")
fila_resumen += 1

ws.cell(row=fila_resumen, column=1, value="Grupo")
ws.cell(row=fila_resumen, column=2, value="Porcentaje")
ws.cell(row=fila_resumen, column=3, value="Cantidad modulos")
ws.cell(row=fila_resumen, column=4, value="Cantidad hs cat")
ws.cell(row=fila_resumen, column=5, value="Cargo")

fila_resumen += 1

if df_resumen_pivot.empty:
    ws.cell(row=fila_resumen, column=1, value="Sin grupos seleccionados")
else:
    for _, row in df_resumen_pivot.iterrows():
        ws.cell(row=fila_resumen, column=1, value=valor_excel_seguro(row["GRUPO"]))
        ws.cell(row=fila_resumen, column=2, value=valor_excel_seguro(row["PORCENTAJE"]))
        ws.cell(row=fila_resumen, column=3, value=valor_excel_seguro(row["CANTIDAD_MODULOS"]))
        ws.cell(row=fila_resumen, column=4, value=valor_excel_seguro(row["CANTIDAD_HS_CAT"]))
        ws.cell(row=fila_resumen, column=5, value=valor_excel_seguro(row["CARGO"]))
        fila_resumen += 1

for column_cells in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column_cells[0].column)

    for cell in column_cells:
        try:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        except Exception:
            pass

    adjusted_width = min(max_length + 2, 60)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(buffer)
buffer.seek(0)

st.download_button(
    label="Exportar vista actual (36 meses)",
    data=buffer.getvalue(),
    file_name=f"rango_36_meses_{fecha_inicio.strftime('%d%m%Y')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.caption(
    f"Exportando cargos con actividad real entre "
    f"{fecha_inicio:%d/%m/%Y} -> {fecha_fin:%d/%m/%Y}"
)